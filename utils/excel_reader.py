"""Excel schema inference -- same output format as infer_csv_schema().

Reads .xlsx / .xls / .xlsm files and infers Power BI column types.
Falls back gracefully when optional dependencies (openpyxl, pandas) are absent.

The returned schema dict is identical to what infer_csv_schema() returns so
the existing SchemaAgent / DAXAgent / ReportAgent pipeline works unchanged.
"""

from __future__ import annotations

import csv
import io
import re
from pathlib import Path
from typing import Any


# ---------------------------------------------------------------------------
# helpers (reused from schema_inference)
# ---------------------------------------------------------------------------

_DATE_RE = re.compile(
    r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}$|^\d{1,2}[-/]\d{1,2}[-/]\d{4}$"
)


def _safe_name(raw: str) -> str:
    """Turn a header cell into a TMDL-safe column name."""
    name = re.sub(r"[^A-Za-z0-9_ ]", "", str(raw)).strip()
    if not name:
        name = "Column"
    if name[0].isdigit():
        name = "C_" + name
    return name


def _infer_pb_type(values: list[Any]) -> str:
    """Guess the Power BI dataType from a sample of cell values."""
    non_null = [v for v in values if v is not None and str(v).strip() != ""]
    if not non_null:
        return "string"

    # Check for boolean
    bool_vals = {str(v).strip().lower() for v in non_null}
    if bool_vals.issubset({"true", "false", "yes", "no", "1", "0"}):
        return "boolean"

    # Try numeric
    int_count = 0
    float_count = 0
    for v in non_null:
        try:
            f = float(str(v))
            if f == int(f):
                int_count += 1
            else:
                float_count += 1
        except (ValueError, TypeError):
            pass

    total = len(non_null)
    if int_count + float_count >= 0.9 * total:
        return "int64" if float_count == 0 else "double"

    # Check date
    date_count = sum(1 for v in non_null if _DATE_RE.match(str(v).strip()))
    if date_count >= 0.9 * total:
        return "dateTime"

    return "string"


def _default_summarize(pb_type: str) -> str:
    if pb_type in {"int64", "double", "decimal"}:
        return "sum"
    return "none"


def _build_schema(table_name: str, headers: list[str],
                  rows: list[list[Any]], source_file: str) -> dict[str, Any]:
    """Build a schema dict from headers + row data."""
    safe_headers = [_safe_name(h) for h in headers]
    columns = []
    for idx, name in enumerate(safe_headers):
        sample = [row[idx] for row in rows if idx < len(row)][:100]
        pb_type = _infer_pb_type(sample)
        columns.append({
            "name": name,
            "dataType": pb_type,
            "summarizeBy": _default_summarize(pb_type),
            "sourceColumn": name,
            "sample_values": [str(v) for v in sample[:5] if v is not None],
        })
    return {
        "table_name": table_name,
        "row_count": len(rows),
        "columns": columns,
        "inferred_relationships": [],
        "source_file": source_file,
        "inferer": "excel",
    }


# ---------------------------------------------------------------------------
# openpyxl path (preferred)
# ---------------------------------------------------------------------------

def _read_with_openpyxl(path: Path, sheet_name: str | None) -> list[dict[str, Any]]:
    """Read Excel using openpyxl; returns list of schema dicts (one per sheet)."""
    import openpyxl  # type: ignore
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    if sheet_name:
        sheets = [sheet_name]
    else:
        sheets = wb.sheetnames

    results = []
    for sname in sheets:
        if sname not in wb.sheetnames:
            raise ValueError(f"Sheet '{sname}' not found. Available: {wb.sheetnames}")
        ws = wb[sname]
        rows_iter = list(ws.iter_rows(values_only=True))
        if not rows_iter:
            continue
        headers = [str(c) if c is not None else f"Col{i}" for i, c in enumerate(rows_iter[0])]
        data_rows = [list(r) for r in rows_iter[1:]]
        tbl = re.sub(r"[^A-Za-z0-9_]", "_", sname).strip("_") or "Sheet"
        results.append(_build_schema(tbl, headers, data_rows, str(path)))

    wb.close()
    return results


# ---------------------------------------------------------------------------
# pandas path (fallback)
# ---------------------------------------------------------------------------

def _read_with_pandas(path: Path, sheet_name: str | None) -> list[dict[str, Any]]:
    """Read Excel using pandas; returns list of schema dicts."""
    import pandas as pd  # type: ignore
    target = sheet_name or 0  # 0 = first sheet if not specified
    df_map = pd.read_excel(path, sheet_name=target)

    if isinstance(df_map, dict):
        items = df_map.items()
    else:
        items = [(path.stem if sheet_name is None else sheet_name, df_map)]

    results = []
    for sname, df in items:
        headers = [str(c) for c in df.columns]
        rows = df.values.tolist()
        tbl = re.sub(r"[^A-Za-z0-9_]", "_", str(sname)).strip("_") or "Sheet"
        results.append(_build_schema(tbl, headers, rows, str(path)))
    return results


# ---------------------------------------------------------------------------
# public API
# ---------------------------------------------------------------------------

def infer_excel_schema(path: Path,
                       sheet_name: str | None = None) -> dict[str, Any]:
    """Read an Excel file and return a schema dict matching infer_csv_schema().

    If the file has multiple sheets and no sheet_name is given, returns the
    schema of the first sheet. Call ``infer_all_sheets()`` for multi-table.

    Args:
        path:       Absolute path to .xlsx / .xls / .xlsm file.
        sheet_name: Sheet name to read; None = first sheet.

    Returns:
        Schema dict with table_name, columns, row_count, etc.

    Raises:
        FileNotFoundError: if path does not exist.
        ImportError: if neither openpyxl nor pandas is installed.
    """
    path = Path(path).expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError(f"Excel file not found: {path}")

    schemas: list[dict[str, Any]] = []

    try:
        schemas = _read_with_openpyxl(path, sheet_name)
    except ImportError:
        pass

    if not schemas:
        try:
            schemas = _read_with_pandas(path, sheet_name)
        except ImportError:
            raise ImportError(
                "Reading Excel files requires either 'openpyxl' or 'pandas'. "
                "Install with: pip install openpyxl"
            )

    if not schemas:
        raise ValueError(f"No data found in '{path}'")

    # return first (or only) schema
    return schemas[0]


def infer_all_sheets(path: Path) -> list[dict[str, Any]]:
    """Return a schema dict for every non-empty sheet in an Excel file."""
    path = Path(path).expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError(f"Excel file not found: {path}")

    try:
        return _read_with_openpyxl(path, sheet_name=None)
    except ImportError:
        pass

    try:
        return _read_with_pandas(path, sheet_name=None)
    except ImportError:
        raise ImportError("Install openpyxl or pandas to read Excel files.")
